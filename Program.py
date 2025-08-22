import asyncio
import aiohttp
import requests
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import os
import time
import json
import threading
import shutil
from openpyxl import load_workbook
import sys
import win32com.client
import re
import winsound
import webbrowser

# Глобальные переменные
base_url, app_token, user_token = '', '', ''
headers = {}
field_mappings = {
    'Тип': 'type',
    'Инв. номер': 'otherserial',
    'Наименование': 'name',
    'Департамент': 'groups_id',
    'Статус': 'states_id',
    'Стеллаж': 'contact',
    'Местоположение': 'locations_id',
    'Пользователь': 'users_id',
    'Комментарий': 'comment'
}
field_visibility = {k: True for k in field_mappings}
entity_display = {
    'User': ('realname', 'firstname', ''),
    'Group': ('name', '', ''),
    'Location': ('completename', '', ''),
    'State': ('completename', '', ''),
    'Contact': ('contact', '', '')
}
use_indexing = True
show_logs = True
debug_mode = False
sound_enabled = True
print_copies = 2
font_scale = 100
main_window_size = '1350x600'
extended_window_size = '1750x700'
auth_window_size = '500x520'
serial_index, entity_index = {}, {}
auth_history = []
auth_success = False

if hasattr(sys, '_MEIPASS'):
    config_file = os.path.join(os.path.dirname(sys.executable), 'config.json')
else:
    config_file = os.path.join(os.path.abspath("."), 'config.json')

if os.path.exists(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            base_url = config.get('base_url', '')
            app_token = config.get('app_token', '')
            user_token = config.get('user_token', '')
            field_mappings = config.get('field_mappings', field_mappings)
            field_visibility = config.get('field_visibility', field_visibility)
            entity_display = config.get('entity_display', entity_display)
            use_indexing = config.get('use_indexing', True)
            show_logs = config.get('show_logs', True)
            debug_mode = config.get('debug_mode', False)
            sound_enabled = config.get('sound_enabled', True)
            print_copies = config.get('print_copies', 2)
            font_scale = config.get('font_scale', 100)
            main_window_size = config.get('main_window_size', '1350x600')
            extended_window_size = config.get('extended_window_size', '1750x700')
            auth_window_size = config.get('auth_window_size', '500x520')
            auth_history = config.get('auth_history', [])
    except Exception as e:
        print(f"Ошибка чтения config.json: {e}")
else:
    auth_history = [
        {'name': 'GLPI', 'base_url': 'https://itdb.proitr.ru/apirest.php',
         'app_token': 'qwevvqP43HhhtA7BYawqNCjmwD5o196PzirczJMz', 'user_token': ''},
        {'name': 'Check', 'base_url': 'https://itdb-check.proitr.ru/apirest.php',
         'app_token': 'qwevvqP43HhhtA7BYawqNCjmwD5o196PzirczJMz', 'user_token': ''}
    ]


def save_config():
    config = {
        'base_url': base_url,
        'app_token': app_token,
        'user_token': user_token,
        'field_mappings': field_mappings,
        'field_visibility': field_visibility,
        'entity_display': entity_display,
        'use_indexing': use_indexing,
        'show_logs': show_logs,
        'debug_mode': debug_mode,
        'sound_enabled': sound_enabled,
        'print_copies': print_copies,
        'font_scale': font_scale,
        'main_window_size': main_window_size,
        'extended_window_size': extended_window_size,
        'auth_window_size': auth_window_size,
        'auth_history': auth_history
    }
    try:
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"Конфигурация сохранена в {config_file}")
    except Exception as e:
        print(f"Ошибка сохранения config.json: {e}")


ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def scaled_font(size):
    return int(size * font_scale / 100)


def trim_location(location):
    prefix = "Про Ай-Ти Ресурс > "
    return location[len(prefix):] if location.startswith(prefix) else location


async def fetch(session, url, headers, params):
    async with session.get(url, headers=headers, params=params) as response:
        return await response.json()


async def index_data_async(app):
    global serial_index, entity_index
    app.log("Начало переиндексации данных...")
    serial_index.clear()
    entity_index.clear()
    item_types = {'Computer': 'Компьютеры', 'Monitor': 'Мониторы', 'Peripheral': 'Устройства'}

    async with aiohttp.ClientSession() as session:
        tasks = []
        for t in item_types:
            url = f'{base_url}/{t}'
            tasks.append(fetch(session, url, headers, {'range': '0-9999'}))
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for t, result in zip(item_types, results):
            if isinstance(result, Exception):
                app.log(f"Ошибка индексации {t}: {result}")
                continue
            if not isinstance(result, list):
                app.log(f"Некорректный ответ для {t}: {result}")
                continue
            for item in result:
                if not isinstance(item, dict):
                    app.log(f"Некорректная запись в {t}: {item}")
                    continue
                if s := item.get('otherserial'):
                    serial_index[s.lstrip('0')] = (t, item)
                if s := item.get('serial'):
                    serial_index[s.lstrip('0')] = (t, item)
            app.log(f"Индексация {t} завершена")

        # Индексация поля contact
        entity_index['Contact'] = set()  # Используем set для уникальных значений
        for t in item_types:
            url = f'{base_url}/{t}'
            tasks = [fetch(session, url, headers, {'range': '0-9999'})]
            data = await asyncio.gather(*tasks, return_exceptions=True)
            data = data[0]
            if isinstance(data, Exception):
                app.log(f"Ошибка индексации Contact для {t}: {data}")
                continue
            if not isinstance(data, list):
                app.log(f"Некорректный ответ для Contact в {t}: {data}")
                continue
            for item in data:
                if not isinstance(item, dict):
                    app.log(f"Некорректная запись Contact в {t}: {item}")
                    continue
                if contact := item.get('contact'):
                    entity_index['Contact'].add(contact)
            app.log(f"Индексация Contact для {t} завершена")

        # Индексация остальных сущностей (без AutoUpdateSystem)
        for entity in ['User', 'Group', 'Location', 'State']:
            entity_index[entity] = {}
            start, step = 0, 500
            while True:
                tasks = [fetch(session, f'{base_url}/{entity}', headers, {'range': f'{start}-{start + step - 1}'})]
                data = await asyncio.gather(*tasks, return_exceptions=True)
                data = data[0]
                if isinstance(data, Exception):
                    app.log(f"Ошибка индексации {entity}: {data}")
                    break
                if not isinstance(data, list):
                    app.log(f"Некорректный ответ для {entity}: {data}")
                    break
                for i in data:
                    if 'id' not in i:
                        app.log(f"Отсутствует 'id' в записи {entity}: {i}")
                        continue
                    field1, field2, _ = entity_display[entity] if len(entity_display[entity]) >= 2 else (entity_display[entity][0], '', '')
                    display_value = i.get(field1, 'Не указано')
                    if entity == 'User':
                        display_value = f"{i.get('realname', '')} {i.get('firstname', '')}".strip()
                    entity_index[entity][str(i['id'])] = display_value
                app.log(f"Индексация {entity}: {start}-{start + len(data) - 1}")
                if len(data) < step:
                    break
                start += step
            app.log(f"Индексация {entity} завершена, элементов: {len(entity_index[entity])}")
    app.log("Переиндексация завершена")


def run_async_index(app):
    asyncio.run(index_data_async(app))


class GLPIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GLPI Inventory Manager")
        self.root.geometry(main_window_size)
        self.root.resizable(True, True)
        self.root.configure(fg_color="#2d2d2d")
        self.change_log, self.found_items, self.pending_serials = [], {}, set()
        self.extended_frame = None
        self.buffer_items = {}
        self.setup_ui()
        self.root.bind("<F11>", lambda e: self.root.state('zoomed'))

    def setup_ui(self):
        self.main_frame = ctk.CTkFrame(self.root, corner_radius=0, fg_color="#2d2d2d")
        self.main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        self.left_frame = ctk.CTkFrame(self.main_frame, corner_radius=10, fg_color="#2d2d2d")
        self.left_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))

        input_frame = ctk.CTkFrame(self.left_frame, corner_radius=10, fg_color="#2d2d2d")
        input_frame.pack(pady=10, fill='x')
        ctk.CTkLabel(input_frame, text="Инв. номер:", font=("Arial", scaled_font(12)), text_color="white").pack(
            side='left', padx=5)
        self.entry = ctk.CTkEntry(input_frame, width=250, font=("Arial", scaled_font(12)))
        self.entry.pack(side='left', padx=5)
        self.entry.bind('<Control-v>', self.paste_handler)
        self.entry.bind('<Return>', lambda e: self.add_serial())
        ctk.CTkButton(input_frame, text="Добавить", command=self.add_serial, font=("Arial", scaled_font(12))).pack(
            side='left', padx=5)

        counter_frame = ctk.CTkFrame(input_frame, fg_color="#2d2d2d")
        counter_frame.pack(side='left', padx=20)
        self.buffer_count_label = ctk.CTkLabel(counter_frame, text="В буфере: 0", font=("Arial", scaled_font(12)),
                                               text_color="white")
        self.buffer_count_label.pack(anchor='w')
        self.found_count_label = ctk.CTkLabel(counter_frame, text="Найдено: 0", font=("Arial", scaled_font(12)),
                                              text_color="white")
        self.found_count_label.pack(anchor='w')

        self.buffer_frame = ctk.CTkFrame(self.left_frame, corner_radius=10, fg_color="#2d2d2d")
        self.buffer_frame.pack(pady=10, fill='both', expand=True)
        self.buffer_canvas = ctk.CTkCanvas(self.buffer_frame, bg="#2d2d2d", highlightthickness=0)
        self.buffer_scrollbar = ctk.CTkScrollbar(self.buffer_frame, command=self.buffer_canvas.yview)
        self.buffer_inner_frame = ctk.CTkFrame(self.buffer_canvas, corner_radius=0, fg_color="#2d2d2d")
        self.buffer_canvas.configure(yscrollcommand=self.buffer_scrollbar.set)
        self.buffer_scrollbar.pack(side='right', fill='y')
        self.buffer_canvas.pack(side='left', fill='both', expand=True)
        self.buffer_canvas.create_window((0, 0), window=self.buffer_inner_frame, anchor="nw")
        self.buffer_inner_frame.bind("<Configure>", lambda e: self.buffer_canvas.configure(
            scrollregion=self.buffer_canvas.bbox("all")))
        self.buffer_canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.buffer_canvas.bind_all('<MouseWheel>', self._on_mousewheel)
        self.buffer_inner_frame.bind('<Enter>', lambda e: self.buffer_canvas.bind_all('<MouseWheel>', self._on_mousewheel))
        self.buffer_inner_frame.bind('<Leave>', lambda e: self.buffer_canvas.unbind_all('<MouseWheel>'))

        button_frame = ctk.CTkFrame(self.left_frame, corner_radius=10, fg_color="#2d2d2d")
        button_frame.pack(pady=10, fill='x')
        ctk.CTkButton(button_frame, text="Очистить", command=self.clear_buffer, font=("Arial", scaled_font(12))).pack(
            side='left', padx=5)
        ctk.CTkButton(button_frame, text="Убрать лишние", command=self.remove_non_green,
                      font=("Arial", scaled_font(12))).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Изменить", command=self.edit_items, font=("Arial", scaled_font(12))).pack(
            side='left', padx=5)
        ctk.CTkButton(button_frame, text="Акты", command=self.acts, font=("Arial", scaled_font(12))).pack(side='left',
                                                                                                          padx=5)
        ctk.CTkButton(button_frame, text="Настройки", command=self.settings_window,
                      font=("Arial", scaled_font(12))).pack(side='left', padx=5)
        # Добавляем кнопку Импорт/Экспорт
        ctk.CTkButton(button_frame, text="Импорт/Экспорт", command=self.import_export_window,
                      font=("Arial", scaled_font(12))).pack(side='left', padx=5)

        self.log_frame = ctk.CTkFrame(self.main_frame, width=300, corner_radius=10, fg_color="#2d2d2d")
        if show_logs:
            self.log_frame.pack(side='right', fill='y')
        ctk.CTkLabel(self.log_frame, text="Логи:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        self.log_text = ctk.CTkTextbox(self.log_frame, height=300, width=280, font=("Arial", scaled_font(10)),
                                       wrap='word')
        self.log_text.pack(fill='y', expand=True, padx=5, pady=5)
        ctk.CTkButton(self.log_frame, text="Очистить логи", command=lambda: self.log_text.delete("0.0", "end"),
                      font=("Arial", scaled_font(12))).pack(pady=5)

        self.labels, self.status_labels, self.info_buttons, self.remove_buttons = {}, {}, {}, {}

    def _on_mousewheel(self, event):
        self.buffer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def log(self, msg):
        if show_logs:
            self.log_text.insert('end', f"[{time.strftime('%H:%M:%S')}] {msg}\n")
            self.log_text.see('end')
        if debug_mode:
            print(f"DEBUG: {msg}")

    def paste_handler(self, event):
        try:
            event.widget.delete(0, 'end')
            clipboard_content = self.root.clipboard_get()
            event.widget.insert(0, clipboard_content)
        except tk.TclError:
            # Если буфер обмена пуст или недоступен
            pass
        return 'break'

    def play_sound(self, success=True):
        if sound_enabled:
            if success:
                winsound.Beep(1000, 200)
            else:
                winsound.Beep(500, 300)

    def open_item_link(self, serial):
        if serial not in self.found_items:
            self.log(f"Ошибка: {serial} не найден для открытия ссылки")
            return
        t, i = self.found_items[serial]
        item_id = i.get('id')
        base_link = "https://itdb.proitr.ru/front/"
        type_map = {
            'Computer': 'computer.form.php',
            'Monitor': 'monitor.form.php',
            'Peripheral': 'peripheral.form.php'
        }
        if t in type_map:
            url = f"{base_link}{type_map[t]}?id={item_id}"
            webbrowser.open(url)
            self.log(f"Открыта ссылка: {url}")

    def update_counters(self):
        buffer_count = len(self.buffer_items)
        found_count = len(self.found_items)
        self.buffer_count_label.configure(text=f"В буфере: {buffer_count}")
        self.found_count_label.configure(text=f"Найдено: {found_count}")

    def add_serial(self):
        s = self.entry.get().strip().lstrip('0')
        if not s:
            self.play_sound(False)
            self.log("Ошибка: пустой ввод")
            return

        buffer_key = f"{s}_{len(self.buffer_items)}"
        self.buffer_items[buffer_key] = s

        row_frame = ctk.CTkFrame(self.buffer_inner_frame, corner_radius=5, fg_color="#2d2d2d")
        row_frame.pack(fill='x', pady=2, padx=5)
        # Прокрутка мышью: только bind на <MouseWheel> для row_frame
        row_frame.bind('<MouseWheel>', lambda e: self.buffer_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
        # --- Отображение: номер, департамент, стеллаж, кнопки ---
        # Инвентарный номер (без департамента, без лишнего width)
        dep_name = ""
        found_item = self.found_items.get(s)
        if found_item:
            t, i = found_item
            dep_id = i.get('groups_id')
            if dep_id and use_indexing and 'Group' in entity_index:
                dep_name = entity_index['Group'].get(str(dep_id), '')
        label_text = s
        if dep_name and dep_name != '-' and dep_name.strip():
            label_text += f" | {dep_name}"
        self.labels[buffer_key] = ctk.CTkLabel(row_frame, text=label_text, font=("Arial", scaled_font(12), "bold"),
                                               text_color="white")
        self.labels[buffer_key].pack(side='left', padx=5)
        # Статус/поиск
        self.status_labels[buffer_key] = ctk.CTkLabel(row_frame, text="Поиск...", fg_color="#ffff00",
                                                      text_color="black", font=("Arial", scaled_font(12)), width=200,
                                                      height=30, corner_radius=5)
        self.status_labels[buffer_key].pack(side='left', padx=5)
        self.status_labels[buffer_key].bind("<Button-1>", lambda e, key=buffer_key: self.open_item_link(
            self.buffer_items.get(key, s)))
        # Кнопки
        self.info_buttons[buffer_key] = ctk.CTkButton(row_frame, text="Инфо",
                                                      command=lambda key=buffer_key: self.show_info(
                                                          self.buffer_items.get(key, s)), state="disabled",
                                                      font=("Arial", scaled_font(12)), width=70)
        self.info_buttons[buffer_key].pack(side='left', padx=5)
        self.remove_buttons[buffer_key] = ctk.CTkButton(row_frame, text="✕",
                                                        command=lambda x=buffer_key: self.remove_serial(x),
                                                        fg_color="#101010", font=("Arial", scaled_font(12)), width=30)
        self.remove_buttons[buffer_key].pack(side='left', padx=5)

        threading.Thread(target=self._search_serial, args=(s, buffer_key), daemon=True).start()
        self.entry.delete(0, 'end')
        self.update_counters()

    def _search_serial(self, s, buffer_key):
        items = []
        for t in ['Computer', 'Monitor', 'Peripheral']:
            try:
                r = requests.get(f'{base_url}/{t}', headers=headers, params={'range': '0-9999'}, timeout=10)
                r.raise_for_status()
                response_data = r.json()
                if not isinstance(response_data, list):
                    self.log(f"Некорректный ответ для {t}: {response_data}")
                    continue
                for item in response_data:
                    otherserial = item.get('otherserial', '').lstrip('0')
                    serial = item.get('serial', '').lstrip('0')
                    # Проверяем совпадение как по инвентарному, так и по серийному номеру
                    if (otherserial and s in otherserial) or (serial and s in serial):
                        items.append((t, item))
            except requests.RequestException as e:
                self.log(f"Ошибка поиска {s} в {t}: {e}")
            except (ValueError, KeyError) as e:
                self.log(f"Ошибка обработки данных для {t}: {e}")

        if not items:
            self.root.after(0, lambda: self.status_labels[buffer_key].configure(fg_color="#992020", text="Не найден",
                                                                                text_color="white"))
            self.play_sound(False)
            self.log(f"Не найден: {s}")
            self.root.after(0, self.update_counters)
        elif len(items) == 1:
            self._process_single_item(s, items[0], buffer_key)
        else:
            self.root.after(0, lambda: self.show_selection_window(s, items, buffer_key))

    def _process_single_item(self, s, item, buffer_key):
        t, i = item
        otherserial = i.get('otherserial', '').lstrip('0')
        serial = i.get('serial', '').lstrip('0')
        # Используем otherserial как ключ, если он есть, иначе serial
        key_serial = otherserial if otherserial else serial
        n = i.get('name', 'Без имени')
        d = n if len(n) <= 30 else n[:27] + '...'
        g = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}[t]

        duplicate = key_serial in self.found_items
        self.buffer_items[buffer_key] = key_serial
        # Обновляем label: добавляем департамент, если найден
        if buffer_key in self.labels:
            dep_name = ""
            dep_id = i.get('groups_id')
            if dep_id and use_indexing and 'Group' in entity_index:
                dep_name = entity_index['Group'].get(str(dep_id), '')
            label_text = key_serial
            if dep_name and dep_name != '-' and dep_name.strip():
                label_text += f" | {dep_name}"
            self.labels[buffer_key].configure(text=label_text)

        if duplicate:
            self.root.after(0, lambda: [
                self.status_labels[buffer_key].configure(fg_color="#101010", text="Дубликат", text_color="white"),
                self.update_counters()
            ])
            self.play_sound(False)
            self.log(f"Дубликат: {key_serial}")
        else:
            self.root.after(0, lambda: [
                self.found_items.update({key_serial: (t, i)}),
                self.status_labels[buffer_key].configure(fg_color="#388938", text=f"{g} {d}", text_color="white"),
                self.info_buttons[buffer_key].configure(state="normal"),
                self.update_counters()
            ])
            self.play_sound(True)
            self.log(f"Найден: {key_serial} ({g})")

    def show_selection_window(self, serial, items, buffer_key):
        win = ctk.CTkToplevel(self.root)
        win.title(f"Выбор для {serial}")
        win.geometry("600x400")
        win.transient(self.root)
        win.grab_set()

        ctk.CTkLabel(win, text=f"Найдено несколько совпадений для {serial}", font=("Arial", scaled_font(14), "bold"),
                     text_color="white").pack(pady=10)
        scroll_frame = ctk.CTkScrollableFrame(win, fg_color="#2d2d2d", width=560)
        scroll_frame.pack(fill='both', expand=True, padx=10, pady=10)

        for t, i in items:
            item_frame = ctk.CTkFrame(scroll_frame, fg_color="#3a3a3a", corner_radius=5)
            item_frame.pack(fill='x', pady=5)
            name = i.get('name', 'Без имени')
            otherserial = i.get('otherserial', '').lstrip('0')
            type_str = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}[t]
            display_text = f"{type_str}: {name} (Инв. № {otherserial})"
            if len(display_text) > 50:
                display_text = display_text[:47] + '...'
            item_label = ctk.CTkLabel(item_frame, text=display_text, font=("Arial", scaled_font(12)),
                                      text_color="white", width=450)
            item_label.pack(side='left', padx=10)
            select_btn = ctk.CTkButton(item_frame, text="Выбрать",
                                       command=lambda it=(t, i): [self._process_single_item(serial, it, buffer_key),
                                                                  win.destroy()], font=("Arial", scaled_font(12)),
                                       width=100)
            select_btn.pack(side='right', padx=10)

    def remove_serial(self, buffer_key):
        s = self.buffer_items[buffer_key]
        self.labels[buffer_key].master.destroy()
        del self.labels[buffer_key], self.info_buttons[buffer_key], self.remove_buttons[buffer_key], self.status_labels[
            buffer_key]
        del self.buffer_items[buffer_key]
        if s not in self.buffer_items.values() and s in self.found_items:
            del self.found_items[s]
        self.log(f"Удалён: {s}")
        self.update_counters()

    def remove_non_green(self):
        if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите убрать все неактуальные позиции?"):
            return
        keys_to_remove = []
        kept_serials = set()

        for buffer_key, s in list(self.buffer_items.items()):
            if s in self.found_items:
                if s not in kept_serials:
                    kept_serials.add(s)
                else:
                    keys_to_remove.append(buffer_key)
            else:
                keys_to_remove.append(buffer_key)

        for buffer_key in keys_to_remove:
            self.remove_serial(buffer_key)

        self.log("Удалены все неактуальные позиции и дубликаты (кроме уникальных зелёных)")

    def show_info(self, s):
        if s not in self.found_items:
            self.play_sound(False)
            self.log("Ошибка: номер не найден для отображения информации")
            return
        self._collapse_extended_frame()
        self.root.geometry(extended_window_size)
        self.extended_frame = ctk.CTkFrame(self.main_frame, width=400, corner_radius=10, fg_color="#2d2d2d")
        self.extended_frame.pack(side='right', fill='both', padx=10)

        t, i = self.found_items[s]
        item_id = i.get('id')

        try:
            r = requests.get(f'{base_url}/{t}/{item_id}', headers=headers, timeout=5)
            r.raise_for_status()
            response_data = r.json()
            if isinstance(response_data, dict):
                i = response_data
                self.found_items[s] = (t, i)
                self.log(f"Информация загружена для {s} (ID: {item_id})")
            else:
                self.log(f"Некорректный ответ для {s}: {response_data}")
        except requests.RequestException as e:
            self.log(f"Ошибка загрузки данных для {s}: {e}")
            i = self.found_items[s][1]
        except (ValueError, KeyError) as e:
            self.log(f"Ошибка обработки данных для {s}: {e}")
            i = self.found_items[s][1]

        ctk.CTkLabel(self.extended_frame, text=f"Инфо: {s}", font=("Arial", scaled_font(14), "bold"),
                     text_color="white", anchor="w").pack(pady=10, anchor="nw", padx=10)

        info_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#3a3a3a")
        info_frame.pack(fill='both', expand=True, pady=10, padx=10)
        info_frame.grid_columnconfigure(1, weight=1)

        e = {'users_id': 'User', 'groups_id': 'Group', 'locations_id': 'Location', 'states_id': 'State'}
        row = 0
        for display_name, glpi_field in field_mappings.items():
            if display_name not in field_visibility or not field_visibility[display_name]:
                continue
            if glpi_field == 'type':
                v = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}[t]
            else:
                v = i.get(glpi_field, 'Не указано')
                if glpi_field == 'contact':  # Прямая работа с contact
                    v = v if v else 'Не указано'
                elif glpi_field in e:
                    entity_type = e[glpi_field]
                    if use_indexing and entity_type in entity_index:
                        v = entity_index[entity_type].get(str(v), 'Не указано')
                        if entity_type == 'Location':
                            v = trim_location(v)
                            if len(v) > 30:
                                parts = v.split('>')
                                if len(parts) > 2:
                                    v = f"{parts[0].strip()}>...>{parts[-1].strip()}"
                                elif len(v) > 40:
                                    v = v[:37] + '...'
                    else:
                        try:
                            r = requests.get(f'{base_url}/{entity_type}/{v}', headers=headers, timeout=5)
                            r.raise_for_status()
                            response_data = r.json()
                            if isinstance(response_data, list) and response_data:
                                data = response_data[0]
                            elif isinstance(response_data, dict):
                                data = response_data
                            else:
                                v = 'Не указано (некорректный ответ)'
                                continue
                            fields = entity_display[entity_type]
                            field1 = fields[0]
                            v = data.get(field1, 'Не указано')
                            if entity_type == 'Location':
                                v = trim_location(v)
                        except requests.RequestException as e:
                            v = f'Не указано (ошибка: {e})'
                        except (ValueError, KeyError) as e:
                            v = f'Не указано (ошибка обработки: {e})'

            row_frame = ctk.CTkFrame(info_frame, fg_color="#3a3a3a")
            row_frame.grid(row=row, column=0, columnspan=2, sticky="nw", pady=2)

            if display_name == 'Комментарий':
                # Используем grid внутри row_frame для точного выравнивания
                row_frame.grid_columnconfigure(1, weight=1)
                label = ctk.CTkLabel(row_frame, text=f"{display_name}:", font=("Arial", scaled_font(12)),
                                     text_color="white", width=150, anchor="w")
                label.grid(row=0, column=0, sticky="nw", padx=5)

                value_textbox = ctk.CTkTextbox(row_frame, font=("Arial", scaled_font(12)), text_color="white",
                                               width=230, wrap="word", fg_color="#3a3a3a", border_width=0)
                value_textbox.insert("0.0", v)
                value_textbox.configure(state="disabled")
                value_textbox.grid(row=0, column=1, sticky="nw", padx=5)
                value_textbox.bind("<Button-1>", lambda e, val=v: self.copy_to_clipboard(val))
                value_textbox.bind("<MouseWheel>",
                                   lambda e, tb=value_textbox: tb.yview_scroll(int(-1 * (e.delta / 120)), "units"))
            else:
                ctk.CTkLabel(row_frame, text=f"{display_name}:", font=("Arial", scaled_font(12)), text_color="white",
                             width=150, anchor="w").pack(side='left', padx=5)
                value_label = ctk.CTkLabel(row_frame, text=v, font=("Arial", scaled_font(12)), text_color="white",
                                           anchor="w", wraplength=230)
                value_label.pack(side='left', padx=5)
                value_label.bind("<Button-1>", lambda e, val=v: self.copy_to_clipboard(val))

            row += 1

        ctk.CTkButton(self.extended_frame, text="Закрыть", command=self._collapse_extended_frame,
                      font=("Arial", scaled_font(12))).pack(pady=10, anchor="nw", padx=10)

    def copy_to_clipboard(self, value):
        self.root.clipboard_clear()
        self.root.clipboard_append(value)
        self.log(f"Скопировано в буфер: {value}")

    def _collapse_extended_frame(self):
        if self.extended_frame:
            self.extended_frame.destroy()
            self.extended_frame = None
            self.root.geometry(main_window_size)
            self.log("Расширенная панель закрыта")

    def clear_buffer(self):
        for w in self.buffer_inner_frame.winfo_children():
            w.destroy()
        self.labels.clear(), self.info_buttons.clear(), self.remove_buttons.clear(), self.status_labels.clear()
        self.found_items.clear(), self.pending_serials.clear(), self.buffer_items.clear()
        self.log("Буфер очищен")
        self.update_counters()

    def settings_window(self):
        self._collapse_extended_frame()
        self.root.geometry("1350x700")
        self.extended_frame = ctk.CTkFrame(self.main_frame, width=300, corner_radius=10, fg_color="#2d2d2d")
        self.extended_frame.pack(side='right', fill='both', padx=10)

        ctk.CTkLabel(self.extended_frame, text="Настройки", font=("Arial", scaled_font(16), "bold"),
                     text_color="white").pack(pady=(10, 20))

        settings_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#3a3a3a")
        settings_frame.pack(fill='x', padx=10, pady=10)

        checkbox_frame = ctk.CTkFrame(settings_frame, fg_color="#3a3a3a")
        checkbox_frame.pack(fill='x', padx=10, pady=10)
        indexing_var = tk.BooleanVar(value=use_indexing)
        ctk.CTkCheckBox(checkbox_frame, text="Использовать индексацию", variable=indexing_var,
                        command=lambda: globals().update(use_indexing=indexing_var.get()),
                        font=("Arial", scaled_font(12)), text_color="white").pack(anchor='w', pady=5)
        logs_var = tk.BooleanVar(value=show_logs)
        ctk.CTkCheckBox(checkbox_frame, text="Показывать поле логов", variable=logs_var,
                        command=lambda: self.toggle_logs(logs_var.get()), font=("Arial", scaled_font(12)),
                        text_color="white").pack(anchor='w', pady=5)
        debug_var = tk.BooleanVar(value=debug_mode)
        ctk.CTkCheckBox(checkbox_frame, text="Включить Debug", variable=debug_var,
                        command=lambda: globals().update(debug_mode=debug_var.get()), font=("Arial", scaled_font(12)),
                        text_color="white").pack(anchor='w', pady=5)
        sound_var = tk.BooleanVar(value=sound_enabled)
        ctk.CTkCheckBox(checkbox_frame, text="Звуковые сигналы", variable=sound_var,
                        command=lambda: globals().update(sound_enabled=sound_var.get()),
                        font=("Arial", scaled_font(12)), text_color="white").pack(anchor='w', pady=5)

        size_frame = ctk.CTkFrame(settings_frame, fg_color="#3a3a3a")
        size_frame.pack(fill='x', padx=10, pady=10)

        ctk.CTkLabel(size_frame, text="Размер главного окна (ШxВ):", font=("Arial", scaled_font(12)),
                     text_color="white").grid(row=0, column=0, padx=5, pady=2, sticky='w')
        main_size_entry = ctk.CTkEntry(size_frame, width=100, font=("Arial", scaled_font(12)))
        main_size_entry.grid(row=0, column=1, padx=5, pady=2)
        main_size_entry.insert(0, main_window_size)

        ctk.CTkLabel(size_frame, text="Размер расширенного окна (ШxВ):", font=("Arial", scaled_font(12)),
                     text_color="white").grid(row=1, column=0, padx=5, pady=2, sticky='w')
        extended_size_entry = ctk.CTkEntry(size_frame, width=100, font=("Arial", scaled_font(12)))
        extended_size_entry.grid(row=1, column=1, padx=5, pady=2)
        extended_size_entry.insert(0, extended_window_size)

        ctk.CTkLabel(size_frame, text="Размер окна авторизации (ШxВ):", font=("Arial", scaled_font(12)),
                     text_color="white").grid(row=2, column=0, padx=5, pady=2, sticky='w')
        auth_size_entry = ctk.CTkEntry(size_frame, width=100, font=("Arial", scaled_font(12)))
        auth_size_entry.grid(row=2, column=1, padx=5, pady=2)
        auth_size_entry.insert(0, auth_window_size)

        ctk.CTkLabel(size_frame, text="Масштаб шрифтов (%):", font=("Arial", scaled_font(12)), text_color="white").grid(
            row=3, column=0, padx=5, pady=2, sticky='w')
        scale_entry = ctk.CTkEntry(size_frame, width=100, font=("Arial", scaled_font(12)))
        scale_entry.grid(row=3, column=1, padx=5, pady=2)
        scale_entry.insert(0, str(font_scale))

        ctk.CTkLabel(size_frame, text="Сколько печатать актов:", font=("Arial", scaled_font(12)),
                     text_color="white").grid(row=4, column=0, padx=5, pady=2, sticky='w')
        copies_entry = ctk.CTkEntry(size_frame, width=100, font=("Arial", scaled_font(12)))
        copies_entry.grid(row=4, column=1, padx=5, pady=2)
        copies_entry.insert(0, str(print_copies))

        def apply_sizes():
            global main_window_size, extended_window_size, auth_window_size, font_scale, print_copies
            try:
                main_size = main_size_entry.get().strip()
                extended_size = extended_size_entry.get().strip()
                auth_size = auth_size_entry.get().strip()
                scale = int(scale_entry.get().strip())
                copies = int(copies_entry.get().strip())

                if not all(re.match(r'^\d+x\d+$', s) for s in [main_size, extended_size, auth_size]):
                    raise ValueError("Формат должен быть ШxВ, например, 1350x600")
                for s in [main_size, extended_size, auth_size]:
                    w, h = map(int, s.split('x'))
                    if not (300 <= w <= 3840 and 300 <= h <= 2160):
                        raise ValueError("Размеры должны быть в пределах 300-3840 по ширине и 300-2160 по высоте")
                if not (50 <= scale <= 200):
                    raise ValueError("Масштаб должен быть от 50 до 200")
                if not (1 <= copies <= 10):
                    raise ValueError("Количество копий должно быть от 1 до 10")

                main_window_size, extended_window_size, auth_window_size, font_scale, print_copies = main_size, extended_size, auth_size, scale, copies
                save_config()
                self.log(
                    f"Размеры обновлены: main={main_size}, extended={extended_size}, auth={auth_size}, scale={scale}, copies={copies}")
                messagebox.showinfo("Успех", "Настройки сохранены. Перезапустите программу для применения изменений.")
                # Убираем self._collapse_extended_frame() — окно остаётся открытым
            except ValueError as e:
                self.log(f"Ошибка в apply_sizes: {e}")
                messagebox.showwarning("Ошибка", str(e))

        button_frame1 = ctk.CTkFrame(self.extended_frame, fg_color="#2d2d2d")
        button_frame1.pack(pady=10, fill='x')
        button_frame2 = ctk.CTkFrame(self.extended_frame, fg_color="#2d2d2d")
        button_frame2.pack(pady=10, fill='x')

        def reindex():
            if 'Session-Token' not in headers:
                messagebox.showwarning("Предупреждение", "Сначала выполните авторизацию!")
                return
            threading.Thread(target=run_async_index, args=(self,), daemon=True).start()
            messagebox.showinfo("Успех", "Переиндексация запущена в фоновом режиме!")
            self.log("Переиндексация запущена")

        def exit_to_auth():
            self.root.withdraw()
            self.clear_buffer()
            self.auth_window()

        ctk.CTkButton(button_frame1, text="Индексация", command=reindex, font=("Arial", scaled_font(12)),
                      width=130).pack(side='left', padx=5)
        ctk.CTkButton(button_frame1, text="Применить", command=apply_sizes, font=("Arial", scaled_font(12)),
                      width=130).pack(side='left', padx=5)
        ctk.CTkButton(button_frame2, text="Сохранить", command=lambda: [save_config(), self.log("Настройки сохранены")],
                      font=("Arial", scaled_font(12)), width=130).pack(side='left', padx=5)
        ctk.CTkButton(button_frame2, text="Закрыть", command=self._collapse_extended_frame,
                      font=("Arial", scaled_font(12)), width=130).pack(side='left', padx=5)
        ctk.CTkButton(button_frame2, text="Выход", command=exit_to_auth, fg_color="#ff5555",
                      font=("Arial", scaled_font(12)), width=130).pack(side='left', padx=5)

    def toggle_logs(self, show):
        global show_logs
        show_logs = show
        if show_logs:
            self.log_frame.pack(side='right', fill='y')
        else:
            self.log_frame.pack_forget()

    def edit_items(self):
        if not self.found_items:
            self.play_sound(False)
            self.log("Ошибка: нет данных")
            return
        self._collapse_extended_frame()
        self.root.geometry(extended_window_size)
        self.extended_frame = ctk.CTkFrame(self.main_frame, width=400, corner_radius=10, fg_color="#2d2d2d")
        self.extended_frame.pack(side='right', fill='both', padx=10)

        ctk.CTkLabel(self.extended_frame, text="Изменение данных", font=("Arial", scaled_font(14), "bold"),
                     text_color="white").pack(pady=10)

        ctk.CTkLabel(self.extended_frame, text="Что будем менять:", font=("Arial", scaled_font(14)),
                     text_color="white").pack(pady=5)
        param_options = ['Пользователь', 'Статус', 'Стеллаж', 'Местоположение', 'Департамент', 'Комментарий']
        param_combo = ctk.CTkComboBox(self.extended_frame, values=param_options, font=("Arial", scaled_font(12)))
        param_combo.pack(pady=5)

        value_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#2d2d2d")
        value_frame.pack(pady=5, fill='x')

        value_widget = None
        user_search_frame = None
        user_list_frame = None
        selected_value = tk.StringVar(value="Выберите значение")
        all_users = []
        letter_combo = None
        number_combo = None
        clear_stelazh_button = None
        apply_timer = None  # Переменная для хранения ID таймера

        def load_users():
            nonlocal all_users
            if use_indexing and 'User' in entity_index:
                all_users = sorted([x for x in entity_index['User'].values() if
                                    isinstance(x, str) and x != 'Не указано' and re.match(r'^[А-Яа-я\s]+$', x)],
                                   key=lambda x: x.lower())
            else:
                try:
                    r = requests.get(f'{base_url}/User', headers=headers, params={'range': '0-9999'}, timeout=10)
                    r.raise_for_status()
                    data = r.json()
                    all_users = sorted([f"{i.get('realname', '')} {i.get('firstname', '')}".strip() for i in data if
                                        (i.get('realname') or i.get('firstname')) and re.match(r'^[А-Яа-я\s]+$',
                                                                                               f"{i.get('realname', '')} {i.get('firstname', '')}".strip())],
                                       key=lambda x: x.lower())
                except requests.RequestException as e:
                    self.log(f"Ошибка загрузки пользователей: {e}")
                    all_users = ['Не указано']
            return all_users

        def update_value_widget(*args):
            nonlocal value_widget, user_search_frame, user_list_frame, letter_combo, number_combo, clear_stelazh_button
            if value_widget:
                value_widget.destroy()
            if user_search_frame:
                user_search_frame.destroy()
            if user_list_frame:
                user_list_frame.destroy()
            if clear_stelazh_button:
                clear_stelazh_button.destroy()
                clear_stelazh_button = None

            selected_value.set("Выберите значение")
            param = param_combo.get()
            if param == 'Пользователь':
                user_search_frame = ctk.CTkFrame(value_frame, corner_radius=10, fg_color="#2d2d2d")
                user_search_frame.pack(fill='x', pady=5)
                ctk.CTkLabel(user_search_frame, text="Поиск пользователя:", font=("Arial", scaled_font(12)),
                             text_color="white").pack(pady=5)
                search_entry = ctk.CTkEntry(user_search_frame, width=300, font=("Arial", scaled_font(12)))
                search_entry.pack(pady=5)

                user_list_frame = ctk.CTkFrame(value_frame, corner_radius=10, fg_color="#2d2d2d")
                user_list_frame.pack(fill='x', pady=5)
                value_widget = ctk.CTkLabel(user_list_frame, textvariable=selected_value,
                                            font=("Arial", scaled_font(12)), text_color="white")
                value_widget.pack(pady=5)

                users = load_users()
                users.insert(0, "Очистить")

                def filter_users(event):
                    query = search_entry.get().strip().lower()
                    filtered = [u for u in users if query in u.lower()] if query else users[:11]
                    for widget in user_list_frame.winfo_children():
                        widget.destroy()
                    value_widget = ctk.CTkLabel(user_list_frame, textvariable=selected_value,
                                                font=("Arial", scaled_font(12)), text_color="white")
                    value_widget.pack(pady=5)
                    if not filtered:
                        ctk.CTkLabel(user_list_frame, text="Нет совпадений", font=("Arial", scaled_font(12)),
                                     text_color="white").pack(pady=5)
                    else:
                        for user in filtered[:11]:
                            ctk.CTkButton(user_list_frame, text=user, command=lambda u=user: selected_value.set(u),
                                          font=("Arial", scaled_font(12)), width=280).pack(pady=2)

                search_entry.bind("<KeyRelease>", filter_users)
                filter_users(None)
            elif param == 'Стеллаж':
                value_widget = ctk.CTkFrame(value_frame, fg_color="#2d2d2d")
                value_widget.pack(pady=5)

                letters = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
                numbers = [str(i) for i in range(6)]
                letter_combo = ctk.CTkComboBox(value_widget, values=letters, font=("Arial", scaled_font(12)), width=100)
                letter_combo.pack(side='left', padx=5)
                number_combo = ctk.CTkComboBox(value_widget, values=numbers, font=("Arial", scaled_font(12)), width=100)
                number_combo.pack(side='left', padx=5)

                def update_selected_value(*args):
                    selected_value.set(f"{letter_combo.get()}{number_combo.get()}")

                letter_combo.configure(command=update_selected_value)
                number_combo.configure(command=update_selected_value)
                letter_combo.set('A')
                number_combo.set('0')
                update_selected_value()

                clear_stelazh_button = ctk.CTkButton(self.extended_frame, text="Очистить стеллаж",
                                                     command=lambda: selected_value.set("Очистить"),
                                                     font=("Arial", scaled_font(12)))
                clear_stelazh_button.pack(pady=5)
            elif param == 'Комментарий':
                value_widget = ctk.CTkEntry(value_frame, width=300, font=("Arial", scaled_font(12)),
                                            placeholder_text="Оставьте пустым для очистки")
                value_widget.pack()
            else:
                value_widget = ctk.CTkComboBox(value_frame, width=300, font=("Arial", scaled_font(12)))
                values = []
                if param == 'Статус':
                    if use_indexing and 'State' in entity_index:
                        values = sorted(
                            [x for x in entity_index['State'].values() if isinstance(x, str) and x != 'Не указано'],
                            key=lambda x: x.lower())
                    else:
                        try:
                            r = requests.get(f'{base_url}/State', headers=headers, params={'range': '0-9999'},
                                             timeout=10)
                            r.raise_for_status()
                            data = r.json()
                            values = sorted([i.get('completename', 'Не указано') for i in data],
                                            key=lambda x: x.lower())
                        except requests.RequestException as e:
                            self.log(f"Ошибка загрузки статусов: {e}")
                            values = ['Не указано']
                elif param == 'Местоположение':
                    if use_indexing and 'Location' in entity_index:
                        values = [trim_location(x) for x in entity_index['Location'].values() if
                                  isinstance(x, str) and x != 'Не указано']
                        shortened = [f"{v.split('>')[0].strip()}>...>{v.split('>')[-1].strip()}" if len(
                            v) > 30 and '>' in v else v[:37] + '...' if len(v) > 40 else v for v in values]
                        values = sorted(shortened, key=lambda x: x.lower())
                    else:
                        try:
                            r = requests.get(f'{base_url}/Location', headers=headers, params={'range': '0-9999'},
                                             timeout=10)
                            r.raise_for_status()
                            data = r.json()
                            values = sorted([trim_location(i.get('completename', 'Не указано')) for i in data],
                                            key=lambda x: x.lower())
                        except requests.RequestException as e:
                            self.log(f"Ошибка загрузки местоположений: {e}")
                            values = ['Не указано']
                elif param == 'Департамент':
                    if use_indexing and 'Group' in entity_index:
                        values = sorted(
                            [x for x in entity_index['Group'].values() if isinstance(x, str) and x != 'Не указано'],
                            key=lambda x: x.lower())
                    else:
                        try:
                            r = requests.get(f'{base_url}/Group', headers=headers, params={'range': '0-9999'},
                                             timeout=10)
                            r.raise_for_status()
                            data = r.json()
                            values = sorted([i.get('name', 'Не указано') for i in data], key=lambda x: x.lower())
                        except requests.RequestException as e:
                            self.log(f"Ошибка загрузки департаментов: {e}")
                            values = ['Не указано']
                values.insert(0, "Очистить")
                value_widget.configure(values=values)
                value_widget.set(values[0])
                value_widget.pack()

        param_combo.configure(command=update_value_widget)
        update_value_widget()

        def apply():
            nonlocal apply_timer
            global use_indexing, entity_index, base_url, headers
            apply_button = apply_btn  # Сохраняем ссылку на кнопку для использования внутри функции

            # Если кнопка уже в состоянии "Уверены?" и нажата повторно
            if apply_button.cget("text") == "Уверены?":
                if apply_timer is not None:
                    self.root.after_cancel(apply_timer)  # Отменяем таймер
                    apply_timer = None
                apply_button.configure(text="Применить")  # Возвращаем исходный текст

                # Применяем изменения
                param = param_combo.get()
                new_value = value_widget.get() if param != 'Пользователь' and param != 'Стеллаж' else selected_value.get()
                if new_value == "Выберите значение":
                    messagebox.showwarning("Предупреждение", "Выберите значение!")
                    return
                update_data = {}
                if param == 'Пользователь':
                    if new_value == "Очистить":
                        update_data['users_id'] = 0
                    else:
                        if use_indexing and 'User' in entity_index:
                            user_id = next(
                                (k for k, v in entity_index['User'].items() if v.startswith(new_value.split('...')[0])),
                                None)
                        else:
                            try:
                                r = requests.get(f'{base_url}/User', headers=headers, params={'range': '0-9999'},
                                                 timeout=10)
                                r.raise_for_status()
                                data = r.json()
                                user_id = next((i['id'] for i in data if
                                                f"{i.get('realname', '')} {i.get('firstname', '')}".strip().startswith(
                                                    new_value.split('...')[0])), None)
                            except requests.RequestException as e:
                                self.log(f"Ошибка поиска пользователя: {e}")
                                return
                        update_data['users_id'] = user_id
                elif param == 'Статус':
                    if new_value == "Очистить":
                        update_data['states_id'] = 0
                    else:
                        if use_indexing and 'State' in entity_index:
                            state_id = next(
                                (k for k, v in entity_index['State'].items() if
                                 v.startswith(new_value.split('...')[0])),
                                None)
                        else:
                            try:
                                r = requests.get(f'{base_url}/State', headers=headers, params={'range': '0-9999'},
                                                 timeout=10)
                                r.raise_for_status()
                                data = r.json()
                                state_id = next((i['id'] for i in data if
                                                 i.get('completename', '').startswith(new_value.split('...')[0])), None)
                            except requests.RequestException as e:
                                self.log(f"Ошибка поиска статуса: {e}")
                                return
                        update_data['states_id'] = state_id
                elif param == 'Стеллаж':
                    if new_value == "Очистить":
                        update_data['contact'] = ""
                    else:
                        update_data['contact'] = new_value
                elif param == 'Местоположение':
                    if new_value == "Очистить":
                        update_data['locations_id'] = 0
                    else:
                        if use_indexing and 'Location' in entity_index:
                            location_id = next((k for k, v in entity_index['Location'].items() if
                                                trim_location(v).startswith(new_value.split('...')[0])), None)
                        else:
                            try:
                                r = requests.get(f'{base_url}/Location', headers=headers, params={'range': '0-9999'},
                                                 timeout=10)
                                r.raise_for_status()
                                data = r.json()
                                location_id = next((i['id'] for i in data if
                                                    trim_location(i.get('completename', '')).startswith(
                                                        new_value.split('...')[0])), None)
                            except requests.RequestException as e:
                                self.log(f"Ошибка поиска местоположения: {e}")
                                return
                        update_data['locations_id'] = location_id
                elif param == 'Департамент':
                    if new_value == "Очистить":
                        update_data['groups_id'] = 0
                    else:
                        if use_indexing and 'Group' in entity_index:
                            group_id = next((k for k, v in entity_index['Group'].items() if v == new_value), None)
                        else:
                            try:
                                r = requests.get(f'{base_url}/Group', headers=headers, params={'range': '0-9999'},
                                                 timeout=10)
                                r.raise_for_status()
                                data = r.json()
                                group_id = next((i['id'] for i in data if i.get('name') == new_value), None)
                            except requests.RequestException as e:
                                self.log(f"Ошибка поиска департамента: {e}")
                                return
                        update_data['groups_id'] = group_id
                elif param == 'Комментарий':
                    update_data['comment'] = new_value if new_value else ""

                for s, (t, i) in self.found_items.items():
                    if 'comment' in update_data and i.get('comment') and new_value:
                        update_data['comment'] = f"{i.get('comment', '')}\n{new_value}"
                    try:
                        r = requests.put(f'{base_url}/{t}/{i["id"]}', headers=headers, json={'input': update_data},
                                         timeout=10)
                        r.raise_for_status()
                        i.update(update_data)
                        self.found_items[s] = (t, i)
                        self.log(f"Обновлено: {s} ({param}: {new_value if new_value else 'Очищено'})")
                    except requests.RequestException as e:
                        self.log(f"Ошибка обновления {s}: {e}")
            else:
                # Первое нажатие: меняем текст кнопки и запускаем таймер
                apply_button.configure(text="Уверены?")
                apply_timer = self.root.after(5000, lambda: apply_button.configure(text="Применить"))  # 5 секунд

        # Создаём кнопку "Применить"
        apply_btn = ctk.CTkButton(self.extended_frame, text="Применить", command=apply, font=("Arial", scaled_font(12)))
        apply_btn.pack(pady=5)
        ctk.CTkButton(self.extended_frame, text="Закрыть", command=self._collapse_extended_frame,
                      font=("Arial", scaled_font(12))).pack(pady=5)
    def acts(self):
        if not self.found_items:
            self.play_sound(False)
            self.log("Ошибка: нет данных")
            return

        self._collapse_extended_frame()
        self.root.geometry(extended_window_size)
        self.extended_frame = ctk.CTkFrame(self.main_frame, width=400, corner_radius=10, fg_color="#2d2d2d")
        self.extended_frame.pack(side='right', fill='both', padx=10)

        ctk.CTkLabel(self.extended_frame, text="Акты", font=("Arial", scaled_font(14), "bold"),
                     text_color="white").pack(pady=10)

        ctk.CTkLabel(self.extended_frame, text="Тип акта:", font=("Arial", scaled_font(12)), text_color="white").pack(
            pady=5)
        act_types = ['Выдача', 'Возврат', 'Выкуп']
        act_combo = ctk.CTkComboBox(self.extended_frame, values=act_types, font=("Arial", scaled_font(12)))
        act_combo.pack(pady=5)
        act_combo._dropdown_menu.bind("<MouseWheel>",
                                      lambda e: act_combo._dropdown_menu.yview_scroll(int(-1 * (e.delta / 120)),
                                                                                      "units"))

        ctk.CTkLabel(self.extended_frame, text="Пользователь:", font=("Arial", scaled_font(12)),
                     text_color="white").pack(pady=5)
        user_search_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#2d2d2d")
        user_search_frame.pack(fill='x', pady=5)
        search_entry = ctk.CTkEntry(user_search_frame, width=300, font=("Arial", scaled_font(12)))
        search_entry.pack(pady=5)

        user_list_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#2d2d2d")
        user_list_frame.pack(fill='x', pady=5)
        selected_user = tk.StringVar(value="Выберите пользователя")
        user_label = ctk.CTkLabel(user_list_frame, textvariable=selected_user, font=("Arial", scaled_font(12)),
                                  text_color="white")
        user_label.pack(pady=5)

        def load_users():
            if use_indexing and 'User' in entity_index:
                users = sorted([x for x in entity_index['User'].values() if
                                isinstance(x, str) and x != 'Не указано' and re.match(r'^[А-Яа-я\s]+$', x)],
                               key=lambda x: x.lower())
            else:
                try:
                    r = requests.get(f'{base_url}/User', headers=headers, params={'range': '0-9999'}, timeout=10)
                    r.raise_for_status()
                    data = r.json()
                    users = sorted([f"{i.get('realname', '')} {i.get('firstname', '')}".strip() for i in data if
                                    (i.get('realname') or i.get('firstname')) and re.match(r'^[А-Яа-я\s]+$',
                                                                                           f"{i.get('realname', '')} {i.get('firstname', '')}".strip())],
                                   key=lambda x: x.lower())
                except requests.RequestException as e:
                    self.log(f"Ошибка загрузки пользователей: {e}")
                    users = ['Не указано']
            return users

        users = load_users()

        def filter_users(event):
            query = search_entry.get().strip().lower()
            filtered = [u for u in users if query in u.lower()] if query else users[:10]
            for widget in user_list_frame.winfo_children():
                widget.destroy()
            user_label = ctk.CTkLabel(user_list_frame, textvariable=selected_user, font=("Arial", scaled_font(12)),
                                      text_color="white")
            user_label.pack(pady=5)
            if not filtered:
                ctk.CTkLabel(user_list_frame, text="Нет совпадений", font=("Arial", scaled_font(12)),
                             text_color="white").pack(pady=5)
            else:
                for user in filtered[:10]:
                    ctk.CTkButton(user_list_frame, text=user, command=lambda u=user: selected_user.set(u),
                                  font=("Arial", scaled_font(12)), width=280).pack(pady=2)

        search_entry.bind("<KeyRelease>", filter_users)
        filter_users(None)

        button_frame = ctk.CTkFrame(self.extended_frame, corner_radius=10, fg_color="#2d2d2d")
        button_frame.pack(pady=10)

        def print_act():
            act_type = act_combo.get()
            user = selected_user.get()
            if not user or user == "Выберите пользователя":
                messagebox.showwarning("Предупреждение", "Выберите пользователя!")
                return
            filename = self.generate_act(act_type, user)
            if filename:
                self.print_with_dialog(filename)

        def save_act():
            act_type = act_combo.get()
            user = selected_user.get()
            if not user or user == "Выберите пользователя":
                messagebox.showwarning("Предупреждение", "Выберите пользователя!")
                return
            filename = self.generate_act(act_type, user)
            if filename:
                save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel files", "*.xlsx")],
                                                         initialfile=f"act_{act_type}_{time.strftime('%Y%m%d_%H%M%S')}")
                if save_path:
                    while not os.path.exists(filename):
                        time.sleep(0.1)
                    shutil.move(filename, save_path)
                    self.log(f"Акт сохранён как: {save_path}")

        ctk.CTkButton(button_frame, text="Печать", command=print_act, font=("Arial", scaled_font(12))).pack(side='left',
                                                                                                            padx=5)
        ctk.CTkButton(button_frame, text="Сохранить как", command=save_act, font=("Arial", scaled_font(12))).pack(
            side='left', padx=5)
        ctk.CTkButton(button_frame, text="Закрыть", command=self._collapse_extended_frame,
                      font=("Arial", scaled_font(12))).pack(side='left', padx=5)

    def generate_act(self, act_type, user):
        try:
            template_map = {'Выдача': 'issuance.xlsx', 'Возврат': 'return.xlsx', 'Выкуп': 'redemption.xlsx'}
            if act_type not in template_map:
                self.log(f"Неизвестный тип акта: {act_type}")
                return None
                
            template_file = resource_path(template_map[act_type])
            if not os.path.exists(template_file):
                self.log(f"Файл шаблона не найден: {template_file}")
                return None
                
            wb = load_workbook(template_file)
            ws = wb.active

            row_start = 6
            for idx, (sn, (_, i)) in enumerate(self.found_items.items(), start=0):
                row = row_start + idx
                ws[f'C{row}'] = i.get('name', 'Не указано')
                ws[f'G{row}' if act_type != 'Выкуп' else f'I{row}'] = i.get('otherserial', sn)
                ws[f'I{row}'] = i.get('serial', sn) if act_type != 'Выкуп' else None

            user_cell = {'Выдача': 'B28', 'Возврат': 'B36', 'Выкуп': 'B34'}[act_type]
            ws[user_cell] = user

            output_file = f'act_{act_type}_{time.strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(output_file)
            self.log(f"Акт {act_type} сформирован: {output_file}")
            return output_file
        except FileNotFoundError as e:
            self.log(f"Файл шаблона не найден: {e}")
            return None
        except PermissionError as e:
            self.log(f"Ошибка доступа к файлу: {e}")
            return None
        except Exception as e:
            self.log(f"Ошибка при формировании акта: {e}")
            return None

    def print_with_dialog(self, filename):
        try:
            if not os.path.exists(filename):
                self.log(f"Файл для печати не найден: {filename}")
                messagebox.showerror("Ошибка", f"Файл не найден: {filename}")
                return
                
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(filename))
            ws = wb.Worksheets[0]
            for _ in range(print_copies):
                ws.PrintOut()
            wb.Close(SaveChanges=False)
            excel.Quit()
            self.log(f"Акт отправлен на печать: {filename} ({print_copies} копий)")
        except FileNotFoundError as e:
            self.log(f"Файл не найден: {e}")
            messagebox.showerror("Ошибка", f"Файл не найден: {e}")
        except PermissionError as e:
            self.log(f"Ошибка доступа к файлу: {e}")
            messagebox.showerror("Ошибка", f"Ошибка доступа к файлу: {e}")
        except Exception as e:
            self.log(f"Ошибка печати: {e}")
            messagebox.showerror("Ошибка", f"Не удалось напечатать акт: {e}")

    def auth_window(self):
        w = ctk.CTkToplevel(self.root)
        w.title("Авторизация")
        w.geometry(auth_window_size)
        w.resizable(False, False)
        w.configure(fg_color="#2d2d2d")
        self.auth_logic(w)
        self.log("Окно авторизации открыто")

    def auth_logic(self, w):
        global auth_success
        auth_success = False
        f = ctk.CTkFrame(w, corner_radius=10, fg_color="#2d2d2d")
        f.pack(pady=20, padx=20, fill='both', expand=True)

        ctk.CTkLabel(f, text="Прессет:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        history_combo = ctk.CTkComboBox(f, values=[h.get('name', f"{h['base_url']} - {h['user_token'][:8]}...") for h in
                                                   auth_history], font=("Arial", scaled_font(12)))
        history_combo.pack(pady=5)
        history_combo.set("")
        history_combo._dropdown_menu.bind("<MouseWheel>",
                                          lambda e: history_combo._dropdown_menu.yview_scroll(int(-1 * (e.delta / 120)),
                                                                                              "units"))

        ctk.CTkLabel(f, text="Название авторизации:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        name_entry = ctk.CTkEntry(f, width=300, font=("Arial", scaled_font(12)))
        name_entry.pack(pady=5)

        ctk.CTkLabel(f, text="URL:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        u = ctk.CTkEntry(f, width=300, font=("Arial", scaled_font(12)))
        u.pack(pady=5)
        u.insert(0, '' if w.winfo_toplevel() == self.root else base_url)

        ctk.CTkLabel(f, text="APP TOKEN:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        a = ctk.CTkEntry(f, width=300, font=("Arial", scaled_font(12)))
        a.pack(pady=5)
        a.insert(0, '' if w.winfo_toplevel() == self.root else app_token)

        ctk.CTkLabel(f, text="USER TOKEN:", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        ut = ctk.CTkEntry(f, width=300, font=("Arial", scaled_font(12)))
        ut.pack(pady=5)
        ut.insert(0, '' if w.winfo_toplevel() == self.root else user_token)

        remember_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(f, text="Запомнить", variable=remember_var, font=("Arial", scaled_font(12)),
                        text_color="white").pack(pady=10)

        button_frame = ctk.CTkFrame(f, fg_color="#2d2d2d")
        button_frame.pack(pady=5)
        close_button = ctk.CTkButton(button_frame, text="Закрыть", command=w.destroy, font=("Arial", scaled_font(12)))
        close_button.pack(side='left', padx=5)
        apply_button = ctk.CTkButton(button_frame, text="Применить", font=("Arial", scaled_font(12)))
        apply_button.pack(side='left', padx=5)

        def load_history(value):
            selection = history_combo.get()
            if not selection:
                return
            for h in auth_history:
                if h.get('name', f"{h['base_url']} - {h['user_token'][:8]}...") == selection:
                    u.delete(0, 'end')
                    u.insert(0, h['base_url'])
                    a.delete(0, 'end')
                    a.insert(0, h['app_token'])
                    ut.delete(0, 'end')
                    ut.insert(0, h['user_token'])
                    name_entry.delete(0, 'end')
                    name_entry.insert(0, h.get('name', ''))
                    break

        history_combo.configure(command=load_history)

        def apply():
            global base_url, app_token, user_token, headers, auth_success
            base_url, app_token, user_token = u.get(), a.get(), ut.get()
            headers = {'App-Token': app_token, 'Authorization': f'user_token {user_token}',
                       'Content-Type': 'application/json'}
            self.log("Попытка авторизации...")
            try:
                r = requests.get(f'{base_url}/initSession', headers=headers, timeout=5)
                r.raise_for_status()
                response_data = r.json()
                if isinstance(response_data, dict) and 'session_token' in response_data:
                    headers['Session-Token'] = response_data['session_token']
                    self.log("Сессия создана успешно")
                    if remember_var.get():
                        auth_entry = {'base_url': base_url, 'app_token': app_token, 'user_token': user_token,
                                      'name': name_entry.get().strip() or f"{base_url} - {user_token[:8]}..."}
                        if auth_entry not in auth_history:
                            auth_history.append(auth_entry)
                        save_config()
                        self.log("Настройки сохранены")
                    auth_success = True
                    w.destroy()
                    self.log("Показ главного окна...")
                    self.root.deiconify()
                    self.root.update()
                    if use_indexing and 'Session-Token' in headers:
                        threading.Thread(target=run_async_index, args=(self,), daemon=True).start()
                        self.log("Индексация запущена в фоновом режиме")
                else:
                    raise ValueError("Некорректный ответ сервера")
            except requests.RequestException as e:
                self.log(f"Ошибка авторизации: {e}")
                apply_button.configure(fg_color="#ff5555")
                messagebox.showerror("Ошибка", f"Не удалось авторизоваться: {e}")
            except (ValueError, KeyError) as e:
                self.log(f"Ошибка обработки ответа сервера: {e}")
                apply_button.configure(fg_color="#ff5555")
                messagebox.showerror("Ошибка", f"Некорректный ответ сервера: {e}")

        apply_button.configure(command=apply)

        def on_closing_auth():
            if not auth_success:
                messagebox.showwarning("Предупреждение", "Необходима успешная авторизация!")
            else:
                w.destroy()

        w.protocol("WM_DELETE_WINDOW", on_closing_auth)

    def import_export_window(self):
        win = ctk.CTkToplevel(self.root)
        win.title("Импорт/Экспорт инвентарных номеров")
        win.geometry("600x500")
        win.transient(self.root)
        win.grab_set()
        
        # Верхняя панель с выбором режима
        mode_frame = ctk.CTkFrame(win, fg_color="#2d2d2d")
        mode_frame.pack(fill='x', pady=5, padx=10)
        mode_var = tk.StringVar(value="import")
        
        def switch_mode(mode):
            mode_var.set(mode)
            if mode == "import":
                title_label.configure(text="Импорт инвентарных номеров")
                import_frame.pack(fill='both', expand=True, pady=5, padx=10)
                export_frame.pack_forget()
            else:
                title_label.configure(text="Экспорт инвентарных номеров")
                export_frame.pack(fill='both', expand=True, pady=5, padx=10)
                import_frame.pack_forget()
        
        title_label = ctk.CTkLabel(win, text="Импорт инвентарных номеров", font=("Arial", scaled_font(14), "bold"), text_color="white")
        title_label.pack(pady=10)
        
        ctk.CTkButton(mode_frame, text="Импорт", command=lambda: switch_mode("import"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(mode_frame, text="Экспорт", command=lambda: switch_mode("export"), width=120).pack(side='left', padx=5)
        
        # Фрейм для импорта
        import_frame = ctk.CTkFrame(win, fg_color="#232323")
        
        # Кнопки для выбора способа импорта
        import_tab_frame = ctk.CTkFrame(import_frame, fg_color="#2d2d2d")
        import_tab_frame.pack(fill='x', pady=5)
        
        import_mode = tk.StringVar(value="excel")
        def show_import_frame(mode):
            import_mode.set(mode)
            for f in [excel_frame, txt_frame, buffer_frame, filters_frame]:
                f.pack_forget()
            if mode == "excel":
                excel_frame.pack(fill='both', expand=True, pady=10)
            elif mode == "txt":
                txt_frame.pack(fill='both', expand=True, pady=10)
            elif mode == "buffer":
                buffer_frame.pack(fill='both', expand=True, pady=10)
            elif mode == "filters":
                filters_frame.pack(fill='both', expand=True, pady=10)
        
        ctk.CTkButton(import_tab_frame, text="Excel", command=lambda: show_import_frame("excel"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(import_tab_frame, text="TXT/CSV", command=lambda: show_import_frame("txt"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(import_tab_frame, text="Буфер/Вставка", command=lambda: show_import_frame("buffer"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(import_tab_frame, text="Фильтры", command=lambda: show_import_frame("filters"), width=120).pack(side='left', padx=5)
        
        # Excel импорт с выбором столбца
        excel_frame = ctk.CTkFrame(import_frame, fg_color="#232323")
        ctk.CTkLabel(excel_frame, text="Выберите Excel файл (.xlsx)", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        
        # Фрейм для выбора столбца
        column_frame = ctk.CTkFrame(excel_frame, fg_color="#232323")
        column_frame.pack(fill='x', pady=5)
        ctk.CTkLabel(column_frame, text="Столбец:", font=("Arial", scaled_font(12)), text_color="white").pack(side='left', padx=5)
        column_var = tk.StringVar(value="A")
        column_combo = ctk.CTkComboBox(column_frame, values=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"], variable=column_var, width=80)
        column_combo.pack(side='left', padx=5)
        
        def import_excel():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active
                numbers = []
                column = column_var.get()
                
                # Получаем все значения из выбранного столбца
                for row in range(1, ws.max_row + 1):
                    cell_value = ws[f'{column}{row}'].value
                    if cell_value and isinstance(cell_value, (str, int)):
                        s = str(cell_value).strip()
                        if s.isdigit():
                            numbers.append(s)
                
                self._import_serials(numbers)
                self.log(f"Импортировано из Excel (столбец {column}): {len(numbers)} номеров")
                win.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка чтения Excel: {e}")
        
        ctk.CTkButton(excel_frame, text="Выбрать файл", command=import_excel).pack(pady=10)
        
        # TXT/CSV импорт
        txt_frame = ctk.CTkFrame(import_frame, fg_color="#232323")
        ctk.CTkLabel(txt_frame, text="Выберите TXT или CSV файл", font=("Arial", scaled_font(12)), text_color="white").pack(pady=10)
        
        def import_txt():
            file_path = filedialog.askopenfilename(filetypes=[("Text/CSV files", "*.txt;*.csv")])
            if not file_path:
                return
            try:
                numbers = []
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        for part in line.replace(';', ',').replace('\t', ',').split(','):
                            s = part.strip()
                            if s.isdigit():
                                numbers.append(s)
                self._import_serials(numbers)
                self.log(f"Импортировано из TXT/CSV: {len(numbers)} номеров")
                win.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка чтения файла: {e}")
        
        ctk.CTkButton(txt_frame, text="Выбрать файл", command=import_txt).pack(pady=10)
        
        # Буфер/ручной ввод
        buffer_frame = ctk.CTkFrame(import_frame, fg_color="#232323")
        ctk.CTkLabel(buffer_frame, text="Вставьте номера через Enter, запятую или пробел", font=("Arial", scaled_font(12)), text_color="white").pack(pady=10)
        text_box = ctk.CTkTextbox(buffer_frame, height=10, font=("Arial", scaled_font(12)), fg_color="#232323", text_color="white")
        text_box.pack(fill='both', expand=True, padx=10, pady=5)
        
        def paste_clipboard():
            try:
                text_box.delete("0.0", "end")
                text_box.insert("0.0", self.root.clipboard_get())
            except Exception:
                pass
        
        ctk.CTkButton(buffer_frame, text="Вставить из буфера", command=paste_clipboard).pack(pady=5)
        
        def import_from_text():
            raw = text_box.get("0.0", "end").replace(';', ',').replace('\t', ',').replace('\n', ',').replace(' ', ',')
            numbers = [s for s in raw.split(',') if s.strip().isdigit()]
            self._import_serials(numbers)
            self.log(f"Импортировано из вставки: {len(numbers)} номеров")
            win.destroy()
        
        ctk.CTkButton(buffer_frame, text="Импортировать", command=import_from_text).pack(pady=10)
        
        # Фильтры импорт
        filters_frame = ctk.CTkFrame(import_frame, fg_color="#232323")
        ctk.CTkLabel(filters_frame, text="Импорт по фильтрам", font=("Arial", scaled_font(12)), text_color="white").pack(pady=5)
        
        # Список фильтров
        filters_list_frame = ctk.CTkFrame(filters_frame, fg_color="#232323")
        filters_list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        filters_list = []
        
        def add_filter():
            filter_frame = ctk.CTkFrame(filters_list_frame, fg_color="#2a2a2a", corner_radius=5)
            filter_frame.pack(fill='x', pady=2, padx=5)
            logic_var = tk.StringVar(value="И")
            logic_combo = ctk.CTkComboBox(filter_frame, values=["И", "ИЛИ"], variable=logic_var, width=60)
            logic_combo.pack(side='left', padx=5, pady=5)
            field_var = tk.StringVar(value="Департамент")
            field_combo = ctk.CTkComboBox(filter_frame, values=list(field_mappings.keys()), variable=field_var, width=120)
            field_combo.pack(side='left', padx=5, pady=5)
            operator_var = tk.StringVar(value="=")
            operator_combo = ctk.CTkComboBox(filter_frame, values=["=", "!=", "Содержит", "Не содержит"], variable=operator_var, width=100)
            operator_combo.pack(side='left', padx=5, pady=5)
            # Значение: по умолчанию Entry, но если '=' и поле с ограниченным набором — ComboBox
            value_var = tk.StringVar()
            value_entry = ctk.CTkEntry(filter_frame, width=150, placeholder_text="Значение", textvariable=value_var)
            value_entry.pack(side='left', padx=5, pady=5)
            value_combo = None
            def update_value_widget(*args):
                nonlocal value_combo
                if value_combo:
                    value_combo.destroy()
                    value_combo = None
                if value_entry:
                    value_entry.pack_forget()
                field = field_var.get()
                op = operator_var.get()
                # Для этих полей показываем выпадающий список
                limited_fields = {
                    'Департамент': lambda: sorted([v for v in entity_index.get('Group', {}).values() if v != 'Не указано'], key=str.lower),
                    'Статус': lambda: sorted([v for v in entity_index.get('State', {}).values() if v != 'Не указано'], key=str.lower),
                    'Местоположение': lambda: sorted([trim_location(v) for v in entity_index.get('Location', {}).values() if v != 'Не указано'], key=str.lower),
                    'Пользователь': lambda: sorted([v for v in entity_index.get('User', {}).values() if v != 'Не указано'], key=str.lower),
                    'Стеллаж': lambda: sorted(list(entity_index.get('Contact', set())), key=str.lower)
                }
                if op == '=' and field in limited_fields:
                    values = limited_fields[field]()
                    if not values:
                        values = ['Не указано']
                    value_combo = ctk.CTkComboBox(filter_frame, values=values, width=150, variable=value_var)
                    value_combo.pack(side='left', padx=5, pady=5)
                else:
                    value_entry.pack(side='left', padx=5, pady=5)
            field_var.trace_add('write', update_value_widget)
            operator_var.trace_add('write', update_value_widget)
            update_value_widget()
            def remove_filter():
                filter_frame.destroy()
                filters_list.remove(filter_data)
            filter_data = {
                'frame': filter_frame,
                'logic': logic_var,
                'field': field_var,
                'operator': operator_var,
                'value': value_var,
                'remove': remove_filter
            }
            filters_list.append(filter_data)
            ctk.CTkButton(filter_frame, text="✕", command=remove_filter, width=30, fg_color="#ff5555").pack(side='left', padx=5, pady=5)
        
        # Кнопки управления фильтрами
        filter_buttons_frame = ctk.CTkFrame(filters_frame, fg_color="#232323")
        filter_buttons_frame.pack(fill='x', padx=10, pady=5)
        
        ctk.CTkButton(filter_buttons_frame, text="Добавить фильтр", command=add_filter).pack(side='left', padx=5)
        
        def import_by_filters():
            if not filters_list:
                messagebox.showwarning("Предупреждение", "Добавьте хотя бы один фильтр!")
                return
            
            try:
                # Собираем все данные из GLPI
                all_items = []
                for item_type in ['Computer', 'Monitor', 'Peripheral']:
                    try:
                        r = requests.get(f'{base_url}/{item_type}', headers=headers, params={'range': '0-9999'}, timeout=10)
                        r.raise_for_status()
                        data = r.json()
                        if isinstance(data, list):
                            for item in data:
                                if isinstance(item, dict) and 'id' in item:
                                    all_items.append((item_type, item))
                    except Exception as e:
                        self.log(f"Ошибка загрузки {item_type}: {e}")
                
                # Применяем фильтры
                filtered_items = []
                for item_type, item in all_items:
                    matches = True
                    for i, filter_data in enumerate(filters_list):
                        logic = filter_data['logic'].get()
                        field_name = filter_data['field'].get()
                        operator = filter_data['operator'].get()
                        filter_value = filter_data['value'].get().strip()
                        
                        if not filter_value:
                            continue
                        
                        # Получаем значение поля
                        glpi_field = field_mappings[field_name]
                        if glpi_field == 'type':
                            item_value = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}.get(item_type, item_type)
                        elif glpi_field == 'otherserial':
                            item_value = item.get('otherserial', '')
                        elif glpi_field == 'users_id' and use_indexing and 'User' in entity_index:
                            user_id = item.get('users_id')
                            item_value = entity_index['User'].get(str(user_id), '') if user_id else ''
                        elif glpi_field == 'groups_id' and use_indexing and 'Group' in entity_index:
                            group_id = item.get('groups_id')
                            item_value = entity_index['Group'].get(str(group_id), '') if group_id else ''
                        elif glpi_field == 'states_id' and use_indexing and 'State' in entity_index:
                            state_id = item.get('states_id')
                            item_value = entity_index['State'].get(str(state_id), '') if state_id else ''
                        elif glpi_field == 'locations_id' and use_indexing and 'Location' in entity_index:
                            location_id = item.get('locations_id')
                            location = entity_index['Location'].get(str(location_id), '') if location_id else ''
                            item_value = trim_location(location)
                        else:
                            item_value = str(item.get(glpi_field, ''))
                        
                        # Применяем оператор
                        field_matches = False
                        if operator == "=":
                            field_matches = str(item_value).lower() == filter_value.lower()
                        elif operator == "!=":
                            field_matches = str(item_value).lower() != filter_value.lower()
                        elif operator == "Содержит":
                            field_matches = filter_value.lower() in str(item_value).lower()
                        elif operator == "Не содержит":
                            field_matches = filter_value.lower() not in str(item_value).lower()
                        
                        # Применяем логику
                        if i == 0:  # Первый фильтр
                            matches = field_matches
                        elif logic == "И":
                            matches = matches and field_matches
                        elif logic == "ИЛИ":
                            matches = matches or field_matches
                    
                    if matches:
                        # Получаем инвентарный номер
                        serial = item.get('otherserial', '').lstrip('0')
                        if serial:
                            filtered_items.append(serial)
                
                # Импортируем найденные номера
                self._import_serials(filtered_items)
                self.log(f"Импортировано по фильтрам: {len(filtered_items)} номеров")
                win.destroy()
                
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка импорта по фильтрам: {e}")
        
        ctk.CTkButton(filter_buttons_frame, text="Импортировать по фильтрам", command=import_by_filters).pack(side='left', padx=5)
        
        # Добавляем первый фильтр по умолчанию
        add_filter()
        
        # Фрейм для экспорта
        export_frame = ctk.CTkFrame(win, fg_color="#232323")
        
        # Кнопки для выбора способа экспорта
        export_tab_frame = ctk.CTkFrame(export_frame, fg_color="#2d2d2d")
        export_tab_frame.pack(fill='x', pady=5)
        
        export_mode = tk.StringVar(value="excel")
        def show_export_frame(mode):
            export_mode.set(mode)
            for f in [export_excel_frame, export_txt_frame, export_buffer_frame]:
                f.pack_forget()
            if mode == "excel":
                export_excel_frame.pack(fill='both', expand=True, pady=10)
            elif mode == "txt":
                export_txt_frame.pack(fill='both', expand=True, pady=10)
            elif mode == "buffer":
                export_buffer_frame.pack(fill='both', expand=True, pady=10)
        
        ctk.CTkButton(export_tab_frame, text="Excel", command=lambda: show_export_frame("excel"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(export_tab_frame, text="TXT/CSV", command=lambda: show_export_frame("txt"), width=120).pack(side='left', padx=5)
        ctk.CTkButton(export_tab_frame, text="Буфер", command=lambda: show_export_frame("buffer"), width=120).pack(side='left', padx=5)
        
        # Excel экспорт
        export_excel_frame = ctk.CTkFrame(export_frame, fg_color="#232323")
        ctk.CTkLabel(export_excel_frame, text="Экспорт в Excel файл", font=("Arial", scaled_font(12)), text_color="white").pack(pady=10)
        
        def export_excel():
            if not self.found_items:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта!")
                return
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Инвентарные номера"
                # Заголовки
                headers = list(field_mappings.keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                # Данные
                for row_idx, (serial, (item_type, item_data)) in enumerate(self.found_items.items(), start=2):
                    for col_idx, display_name in enumerate(headers, 1):
                        glpi_field = field_mappings[display_name]
                        value = None
                        if glpi_field == 'type':
                            value = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}.get(item_type, item_type)
                        elif glpi_field == 'otherserial':
                            value = serial
                        elif glpi_field == 'users_id' and use_indexing and 'User' in entity_index:
                            value = entity_index['User'].get(str(item_data.get('users_id')), 'Не указано')
                        elif glpi_field == 'groups_id' and use_indexing and 'Group' in entity_index:
                            value = entity_index['Group'].get(str(item_data.get('groups_id')), 'Не указано')
                        elif glpi_field == 'states_id' and use_indexing and 'State' in entity_index:
                            value = entity_index['State'].get(str(item_data.get('states_id')), 'Не указано')
                        elif glpi_field == 'locations_id' and use_indexing and 'Location' in entity_index:
                            loc = entity_index['Location'].get(str(item_data.get('locations_id')), 'Не указано')
                            value = trim_location(loc)
                        else:
                            value = item_data.get(glpi_field, 'Не указано')
                        ws.cell(row=row_idx, column=col_idx, value=value)
                wb.save(file_path)
                self.log(f"Экспортировано в Excel: {len(self.found_items)} записей")
                messagebox.showinfo("Успех", f"Экспортировано {len(self.found_items)} записей в {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка экспорта в Excel: {e}")
        
        ctk.CTkButton(export_excel_frame, text="Экспортировать в Excel", command=export_excel).pack(pady=10)
        
        # TXT/CSV экспорт
        export_txt_frame = ctk.CTkFrame(export_frame, fg_color="#232323")
        ctk.CTkLabel(export_txt_frame, text="Экспорт в TXT или CSV файл", font=("Arial", scaled_font(12)), text_color="white").pack(pady=10)
        
        def export_txt():
            if not self.found_items:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта!")
                return
            file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv"), ("Text files", "*.txt")])
            if not file_path:
                return
            try:
                headers = list(field_mappings.keys())
                with open(file_path, 'w', encoding='utf-8', newline='') as f:
                    f.write(";".join(headers) + "\n")
                    for serial, (item_type, item_data) in self.found_items.items():
                        row = []
                        for display_name in headers:
                            glpi_field = field_mappings[display_name]
                            value = None
                            if glpi_field == 'type':
                                value = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}.get(item_type, item_type)
                            elif glpi_field == 'otherserial':
                                value = serial
                            elif glpi_field == 'users_id' and use_indexing and 'User' in entity_index:
                                value = entity_index['User'].get(str(item_data.get('users_id')), 'Не указано')
                            elif glpi_field == 'groups_id' and use_indexing and 'Group' in entity_index:
                                value = entity_index['Group'].get(str(item_data.get('groups_id')), 'Не указано')
                            elif glpi_field == 'states_id' and use_indexing and 'State' in entity_index:
                                value = entity_index['State'].get(str(item_data.get('states_id')), 'Не указано')
                            elif glpi_field == 'locations_id' and use_indexing and 'Location' in entity_index:
                                loc = entity_index['Location'].get(str(item_data.get('locations_id')), 'Не указано')
                                value = trim_location(loc)
                            else:
                                value = item_data.get(glpi_field, 'Не указано')
                            row.append(str(value))
                        f.write(";".join(row) + "\n")
                self.log(f"Экспортировано в TXT/CSV: {len(self.found_items)} записей")
                messagebox.showinfo("Успех", f"Экспортировано {len(self.found_items)} записей в {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка экспорта в TXT/CSV: {e}")
        
        ctk.CTkButton(export_txt_frame, text="Экспортировать в TXT/CSV", command=export_txt).pack(pady=10)
        
        # Экспорт в буфер
        export_buffer_frame = ctk.CTkFrame(export_frame, fg_color="#232323")
        ctk.CTkLabel(export_buffer_frame, text="Копировать в буфер обмена", font=("Arial", scaled_font(12)), text_color="white").pack(pady=10)
        
        def export_to_buffer():
            if not self.found_items:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта!")
                return
            try:
                headers = list(field_mappings.keys())
                buffer_text = "\t".join(headers) + "\n"
                for serial, (item_type, item_data) in self.found_items.items():
                    row = []
                    for display_name in headers:
                        glpi_field = field_mappings[display_name]
                        value = None
                        if glpi_field == 'type':
                            value = {'Computer': 'Компьютер', 'Monitor': 'Монитор', 'Peripheral': 'Устройство'}.get(item_type, item_type)
                        elif glpi_field == 'otherserial':
                            value = serial
                        elif glpi_field == 'users_id' and use_indexing and 'User' in entity_index:
                            value = entity_index['User'].get(str(item_data.get('users_id')), 'Не указано')
                        elif glpi_field == 'groups_id' and use_indexing and 'Group' in entity_index:
                            value = entity_index['Group'].get(str(item_data.get('groups_id')), 'Не указано')
                        elif glpi_field == 'states_id' and use_indexing and 'State' in entity_index:
                            value = entity_index['State'].get(str(item_data.get('states_id')), 'Не указано')
                        elif glpi_field == 'locations_id' and use_indexing and 'Location' in entity_index:
                            loc = entity_index['Location'].get(str(item_data.get('locations_id')), 'Не указано')
                            value = trim_location(loc)
                        else:
                            value = item_data.get(glpi_field, 'Не указано')
                        row.append(str(value))
                    buffer_text += "\t".join(row) + "\n"
                self.root.clipboard_clear()
                self.root.clipboard_append(buffer_text)
                self.log(f"Скопировано в буфер: {len(self.found_items)} записей")
                messagebox.showinfo("Успех", f"Скопировано {len(self.found_items)} записей в буфер обмена")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка копирования в буфер: {e}")
        
        ctk.CTkButton(export_buffer_frame, text="Копировать в буфер", command=export_to_buffer).pack(pady=10)
        
        # Показываем импорт по умолчанию
        show_import_frame("excel")
        import_frame.pack(fill='both', expand=True, pady=5, padx=10)
        
        ctk.CTkButton(win, text="Закрыть", command=win.destroy, font=("Arial", scaled_font(12))).pack(pady=10)
    def _import_serials(self, numbers):
        count = 0
        for s in numbers:
            self.entry.delete(0, 'end')
            self.entry.insert(0, s)
            self.add_serial()
            count += 1
        self.log(f"Добавлено в буфер: {count}")


def start():
    root = ctk.CTk()
    root.withdraw()
    app = GLPIApp(root)
    print("DEBUG: Запуск программы, открытие окна авторизации")
    app.auth_window()
    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(app))
    root.mainloop()


def on_closing(app):
    if 'Session-Token' in headers:
        try:
            requests.get(f'{base_url}/killSession', headers=headers, timeout=5)
            app.log("Сессия завершена")
        except requests.RequestException as e:
            app.log(f"Ошибка завершения сессии: {e}")
        except Exception as e:
            app.log(f"Неожиданная ошибка при завершении сессии: {e}")
    app.root.destroy()
    print("DEBUG: Программа завершена")


if __name__ == "__main__":
    start()